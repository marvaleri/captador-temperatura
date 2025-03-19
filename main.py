from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
import tkinter as tk

def busca_tempo():
    navegador = webdriver.Chrome()
    navegador.get("https://www.accuweather.com/pt/br/s%C3%A3o-paulo/45881/current-weather/45881")

    temp_atual = navegador.find_element(By.XPATH, '/html/body/div/div[7]/div[1]/div[1]/div[2]/div[2]/div[1]/div[1]/div/div').text
    print(f"A temperatura de hoje é de {temp_atual}")

    umidade_atual = navegador.find_element(By.XPATH, '/html/body/div/div[7]/div[1]/div[1]/div[2]/div[3]/div[6]/div[2]').text
    print(f"A umidade do ar é de {umidade_atual}")

    hora_atual = navegador.find_element(By.XPATH, '/html/body/div/div[7]/div[1]/div[1]/div[2]/div[1]/p').text
    print(f"A hora atual é {hora_atual}")

    data_atual = navegador.find_element(By.XPATH, '/html/body/div/div[7]/div[1]/div[1]/div[1]/div').text
    print(f"A data atual é {data_atual.lower()}")

    arquivo_excel = load_workbook("dados.xlsx")

    planilha = arquivo_excel['tempo']

    planilha.cell(row=2, column=1).value = temp_atual
    planilha.cell(row=2, column=2).value = umidade_atual
    planilha.cell(row=2, column=3).value = hora_atual
    planilha.cell(row=2, column=4).value = data_atual.lower()

    arquivo_excel.save("dados.xlsx")

    print("Fim do programa")
    navegador.quit()

root = tk.Tk()
root.title("Clima e informações do tempo")
root.geometry("300x300")

titulo = tk.Label(
    root, text="Atualizar previsão na planilha", font=("Arial", 16)
)

btn_atualizar = tk.Button(
    root, text="Buscar Previsão", font=("Arial", 12), command=busca_tempo
)

btn_encerrar = tk.Button(
    root, text="Encerrar", font=("Arial", 12), command=root.quit
)

titulo.pack(pady=10)
btn_atualizar.pack(pady=10)
btn_encerrar.pack(pady=10)

tk.mainloop()
